#!/usr/bin/env python3
"""
Build product-costs.json từ file giá vốn Excel.

Input : data/cost-source/kho-tong.xlsx
        (sheet đầu: "Dữ liệu vật tư - Kho tổng" — cột A..T)
Output: data/product-costs.json

Mapping: Mã tên gọi (cột B) → Giá nhập VNĐ (cột G), Giá bán (cột H), Trạng thái (cột J).
Case-insensitive lookup, strip whitespace.

Dùng cho dashboard để tính lợi nhuận:
  Lợi nhuận = Giá bán - Giá bán × 10% VAT - Chi phí quảng cáo - Giá nhập × SL
"""
import os
import re
import json
import sys

try:
    import openpyxl
except ImportError:
    print("[FATAL] openpyxl required. Install: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
SRC = os.path.join(ROOT, "data", "cost-source", "kho-tong.xlsx")
OUT = os.path.join(ROOT, "data", "product-costs.json")


def parse_money(v):
    """Parse cells that may contain '₫404,917' or '$91.00' or plain numbers. Returns float or None."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v)
    # Detect currency symbol
    if "$" in s or "¥" in s:
        # Foreign currency – skip (we only want VND giá nhập từ cột G)
        return None
    # Strip VND symbol and commas
    digits = re.sub(r"[^\d.]", "", s)
    return float(digits) if digits else None


def build():
    if not os.path.exists(SRC):
        print(f"[FATAL] Missing source file: {SRC}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb[wb.sheetnames[0]]

    products = {}
    skipped = 0

    for r in range(2, ws.max_row + 1):
        dinh_danh = ws.cell(r, 1).value
        ma = ws.cell(r, 2).value
        sku = ws.cell(r, 3).value
        ten = ws.cell(r, 4).value
        phan_loai = ws.cell(r, 5).value
        gia_nhap_vnd = parse_money(ws.cell(r, 7).value)
        gia_ban = parse_money(ws.cell(r, 8).value)
        don_vi = ws.cell(r, 9).value
        trang_thai = ws.cell(r, 10).value
        ton_kho = ws.cell(r, 11).value

        if not ma:
            skipped += 1
            continue
        ma_clean = str(ma).strip()
        if not ma_clean:
            skipped += 1
            continue

        key = ma_clean.lower()
        products[key] = {
            "ma_ten_goi": ma_clean,
            "dinh_danh": (str(dinh_danh).strip() if dinh_danh else ""),
            "sku": (str(sku).strip() if sku else ""),
            "ten": (str(ten).strip() if ten else ""),
            "phan_loai": (str(phan_loai).strip() if phan_loai else ""),
            "gia_nhap_vnd": gia_nhap_vnd,
            "gia_ban_vnd": gia_ban,
            "don_vi": (str(don_vi).strip() if don_vi else ""),
            "trang_thai": (str(trang_thai).strip() if trang_thai else ""),
            "ton_kho": ton_kho,
            "row": r,
        }

    # Explicit alias map for Pancake code variations → xlsx "Mã tên gọi"
    # Giữ key lowercase cho lookup
    aliases = {
        # DR1 Pancake code = "dr1 new" → xlsx "DR1 New" (đang kinh doanh)
        "dr1": "dr1 new",
        # DA8.1 Pro Pancake → xlsx "DA8.1 PRO" (case khác)
        "da8.1 pro": "da8.1 pro",  # ma_clean là 'DA8.1 PRO', key đã lowercase
    }

    output = {
        "source_file": "data/cost-source/kho-tong.xlsx",
        "row_count": len(products),
        "skipped_rows": skipped,
        "aliases": aliases,  # Pancake-code → xlsx-key (both lowercase)
        "products": products,
    }

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"[OK] Wrote {OUT}")
    print(f"     {len(products)} sản phẩm, {skipped} dòng bỏ qua")

    # Sanity check: 13 SKU profit tracking
    MAIN = [
        "d1", "d1 pro", "d2", "d3", "d4", "d8 pro",
        "dr1 new", "dr4 plus",
        "dv1 pro",
        "da8.1", "da8.1 pro",
        "noma 911", "noma 922",
    ]
    print("\n[CHECK] 13 SKU profit tracking:")
    for m in MAIN:
        p = products.get(m)
        if not p:
            print(f"  ❌ {m!r} không tìm thấy")
            continue
        gn = p["gia_nhap_vnd"]
        gb = p["gia_ban_vnd"]
        tt = p["trang_thai"]
        status = "✓" if gn and gn > 0 else "⚠️ thiếu giá nhập"
        print(f"  {status} {m!r}: giá nhập = {gn:,.0f}đ · giá bán = {gb:,.0f}đ · {tt}" if gn and gb else f"  {status} {m!r}: gn={gn}, gb={gb}, tt={tt}")


if __name__ == "__main__":
    build()
